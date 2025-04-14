import re
from openpyxl import load_workbook

# === CONFIGURATION ===
input_file = "sample_algosec_export.xlsx"          # Path to your input Excel file
output_file = "sample_algosec_export_cleaned.xlsx" # Path to save the cleaned file
input_column_name = "Destination"                  # Column to clean
output_column_name = "Cleaned Destinations"        # Where to write cleaned IPs

# === IP Matching Logic ===
ip_pattern = re.compile(r'\b(?:\d{1,3}\.){3}\d{1,3}\b')

def is_valid_ip(ip):
    try:
        return all(0 <= int(part) <= 255 for part in ip.split('.'))
    except ValueError:
        return False

def extract_valid_ips(raw_text):
    if raw_text is None:
        return []
    raw_text = str(raw_text).replace(';', ',')  # Replace semicolons with commas
    ips = ip_pattern.findall(raw_text)
    return [ip for ip in ips if is_valid_ip(ip)]

# === Load Workbook ===
wb = load_workbook(input_file)
ws = wb.active

# === Find the Target Column ===
header = [cell.value for cell in ws[1]]
try:
    col_index = header.index(input_column_name) + 1  # 1-based index
except ValueError:
    raise Exception(f"Column '{input_column_name}' not found in the header row.")

# === Write Header for Output Column ===
output_col_index = len(header) + 1
ws.cell(row=1, column=output_col_index, value=output_column_name)

# === Process Each Row ===
for row in ws.iter_rows(min_row=2, min_col=col_index, max_col=col_index):
    raw_cell = row[0].value
    cleaned_ips = extract_valid_ips(raw_cell)
    cleaned_text = ', '.join(cleaned_ips)
    ws.cell(row=row[0].row, column=output_col_index, value=cleaned_text)

# === Save the Cleaned File ===
wb.save(output_file)
print(f"Cleaned file saved to: {output_file}")
